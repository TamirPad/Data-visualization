import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime
import json
import calendar


# this module iterates all excel files in a directory and return a key-value dict {'filename': df} of Dataframes
# notice this module is relying that 'assined shifts' data is at J column.

# get a path to excel file and return json obj
def excel_to_json(path):
    # load a work book and worksheet
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # creating shift dictionaries by extracting data from col A, B, C and I on
    # the excel sheet  and appending it to a shifts list
    shiftslist = []
    for row in range(2, 95):
        # formatting the dates
        date_cell = str(ws['B' + str(row)].value)

        if date_cell != 'None':
            DT = datetime.strptime(date_cell, '%Y-%m-%d %H:%M:%S')
            formatted_date = DT.strftime('%Y-%m-%d')

        # Stop Condition whenever hits 'none' on 'employee' --> end of the month
        if str(ws['J' + str(row)].value) == 'None':
            break
        # creating shifts list
        shift = {
            "date": formatted_date,
            "day": calendar.day_name[DT.weekday()],
            "hours": ws['C' + str(row)].value,
            "employee": ws['J' + str(row)].value
        }
        shiftslist.append(shift)

    # creating a json shifts obj and Writing it to a file
    shifts_json = json.dumps(shiftslist)
    return shifts_json


def get_monthly_df():
    # assign directory
    directory = 'excel_shifts'

    monthly_df = {}
    pd.set_option('display.max_columns', 14)
    pd.set_option('display.max_rows', 93)
    # iterate over files in that directory
    for filename in os.listdir(directory):
        f = os.path.join(directory, filename)
        # checking if it is a file
        if os.path.isfile(f):
            monthly_df[f'{filename}'] = pd.read_json(excel_to_json('excel_shifts\\' + filename))

    return monthly_df
